#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
════════════════════════════════════════════════════════════════════
  GÜMRÜK BEYANNAME HAZIRLIK SİSTEMİ
  CI + PL  ->  Tarife bazında beyanname hazırlık Excel'i (+ hata yakalama + öğrenme)
════════════════════════════════════════════════════════════════════

KULLANIM:
  python3 beyanname_hazirla.py --ci CI.xls --pl PL.xls --tedarikci "ENEMOB LIMITED" --dosya AKKU4100766

  Zorunlu:
    --ci        Commercial Invoice dosyası (.xls/.xlsx)
    --pl        Packing List dosyası (.xls/.xlsx)
    --dosya     Dosya/beyanname referans adı (çıktı adı olur)
  Opsiyonel:
    --tedarikci Tedarikçi adı (öğrenme anahtarı; CI'dan otomatik de bulunur)
    --mense     Menşe ülke kodu (varsayılan: öğrenme dosyasından/boş)
    --onayla    Bu bayrak verilirse çıktı öğrenme dosyasına işlenir (SEN ONAYLADIKTAN SONRA)

ÇIKTI (cikti/ klasörüne):
    <dosya>_HAZIRLIK.xlsx   4 sayfa: Detay_Eslestirme, Tarife_Bazinda_Ozet, Kontrol, XML_Hazirlik
"""
import pandas as pd, numpy as np, re, os, sys, argparse, pickle
from datetime import datetime
from difflib import SequenceMatcher
import warnings
warnings.filterwarnings('ignore')

# ---- Yollar (script konumuna göre) ----
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REF_PKL   = os.path.join(BASE, 'veri', 'reference.pkl')
LEARN_XLSX= os.path.join(BASE, 'veri', 'ogrenilmis_eslesmeler.xlsx')
CIKTI_DIR = os.path.join(BASE, 'cikti')

# ════════════════════ YARDIMCI ════════════════════
def norm(s):
    if pd.isna(s) or s is None: return None
    s = re.sub(r'\s+',' ',str(s).strip())
    return s if s else None
def normU(s):
    n = norm(s); return n.upper() if n else None
def dot_gtip(g):
    g=str(g).replace('.','').replace(' ','')
    return f"{g[0:4]}.{g[4:6]}.{g[6:8]}.{g[8:10]}.{g[10:12]}" if len(g)==12 else g
def read_any(path):
    ext=os.path.splitext(path)[1].lower()
    eng='xlrd' if ext=='.xls' else 'openpyxl'
    return pd.read_excel(path, engine=eng, header=None)

# ════════════════════ CI META (satıcı/menşe/fatura no otomatik) ════════════════════
def extract_meta(path):
    """CI'dan tedarikçi (faturayı kesen), menşe, fatura no otomatik çıkar."""
    raw = read_any(path)
    # tüm hücreleri metin olarak topla (ilk 15 satır = başlık bölgesi)
    head_lines=[]
    for i in range(min(20,len(raw))):
        vals=[str(v).strip() for v in raw.iloc[i] if pd.notna(v) and str(v).strip()]
        if vals: head_lines.append(' | '.join(vals))
    head=' \n '.join(head_lines)
    alltext=' '.join(str(v) for i in range(len(raw)) for v in raw.iloc[i] if pd.notna(v))

    # --- Tedarikçi: genelde ilk satır(lar)da, satıcı firma adı (CO/LTD/INC/GMBH...) ---
    tedarikci=None
    for i in range(min(6,len(raw))):
        for v in raw.iloc[i]:
            if pd.isna(v): continue
            t=re.sub(r'\s+',' ',str(v).strip())
            if re.search(r'\b(CO\.?,?\s*LTD|LIMITED|LTD|INC|GMBH|S\.?P\.?A|TECHNOLOGY|ELECTRONICS|INDUSTR|INTERNATIONAL|INVESTMENT|CORPORATION|MANUFACTUR)\b', t, re.I) \
               and not re.search(r'INVOICE|PACKING', t, re.I) and len(t)>5:
                tedarikci=t; break
        if tedarikci: break

    # --- Fatura No ---
    fatno=None
    m=re.search(r'(?:INVOICE\s*NO\.?|CI\s*NO\.?|INV\.?\s*NO\.?)[:\s]*([A-Z0-9\-\/]{5,})', alltext, re.I)
    if m: fatno=m.group(1).strip()
    if not fatno:
        m=re.search(r'\b(CI\d{6,}|HD\d{6,}|[A-Z]{2,}\d{6,})\b', alltext)
        if m: fatno=m.group(1)

    # --- Menşe ---
    mense_kod='720'  # varsayılan Çin
    m=re.search(r'(?:COUNTRY OF ORIGIN|ORIGIN)[:\s]*([A-Za-z ]+)', alltext, re.I)
    if m:
        ulke=m.group(1).strip().upper()
        harita={'CHINA':'720','TURKEY':'052','ITALY':'005','GERMANY':'004','USA':'400',
                'UNITED STATES':'400','SOUTH KOREA':'728','KOREA':'728','JAPAN':'732',
                'VIETNAM':'690','INDIA':'664','TAIWAN':'736'}
        for k,c in harita.items():
            if k in ulke: mense_kod=c; break
    return dict(tedarikci=tedarikci, fatno=fatno, mense_kod=mense_kod)

def kisa_ad(tedarikci):
    """Tedarikçiden dosya adı öneki üret: ilk anlamlı kelime, büyük harf."""
    if not tedarikci: return 'FIRMA'
    t=re.sub(r'[^A-Za-z ]',' ',tedarikci).strip()
    stop={'THE','CO','LTD','LIMITED','INC','GMBH','GUANGZHOU','SHENZHEN','NINGBO','CIXI'}
    for w in t.split():
        if w.upper() not in stop and len(w)>=3:
            return w.upper()[:12]
    return t.split()[0].upper()[:12] if t.split() else 'FIRMA'

# ════════════════════ CI OKUMA ════════════════════
def read_ci(path):
    raw=read_any(path)
    hr=None
    for i in range(min(30,len(raw))):
        t=' '.join(str(v).lower() for v in raw.iloc[i] if pd.notna(v))
        if ('product' in t or 'description' in t or 'ürün' in t) and ('quantity' in t or 'qty' in t or 'adet' in t or 'pcs' in t):
            hr=i;break
    if hr is None: raise SystemExit("HATA: CI başlık satırı bulunamadı (Product+Quantity içeren satır yok).")
    hdr=[str(v).strip().lower() if pd.notna(v) else f'c{j}' for j,v in enumerate(raw.iloc[hr])]
    def col(keys):
        for j,h in enumerate(hdr):
            if any(k in h for k in keys): return j
        return None
    c_item=col(['item']); c_prod=col(['product','description','ürün','goods'])
    c_model=col(['model','part']); c_price=col(['unit price','price','fiyat','birim'])
    c_qty=col(['quantity','qty','adet','pcs']); c_total=col(['total','tutar','amount'])
    c_gtip=col(['remark','gtip','hs code','tarife'])
    # GTİP başlıkla bulunamayabilir veya değer bir sağdaki kolonda olabilir.
    # Veri satırlarını tarayıp GTİP-benzeri (10-12 haneli, noktalı) değer içeren kolonu bul.
    def looks_gtip(v):
        if pd.isna(v): return False
        g=re.sub(r'[^\d]','',str(v))
        return len(g) in (10,11,12)
    gtip_scores={}
    for i in range(hr+1,min(hr+20,len(raw))):
        for j in range(len(raw.columns)):
            if looks_gtip(raw.iloc[i,j]):
                gtip_scores[j]=gtip_scores.get(j,0)+1
    if gtip_scores:
        c_gtip=max(gtip_scores,key=gtip_scores.get)
    rows=[]; last=None
    for i in range(hr+1,len(raw)):
        r=raw.iloc[i]
        it=r[c_item] if c_item is not None else r[0]
        if pd.isna(it) or not str(it).strip().replace('.','').isdigit(): continue
        name=norm(r[c_prod]) if c_prod is not None else None
        if name: last=name
        else: name=last
        gtip=None
        if c_gtip is not None and pd.notna(r[c_gtip]):
            g=re.sub(r'[^\d]','',str(r[c_gtip]))
            gtip=g if len(g)>=8 else None
        rows.append(dict(ci_satir=int(float(it)), urun=name,
            model=norm(r[c_model]) if c_model is not None else None,
            birim_fiyat=r[c_price] if c_price is not None else None,
            adet_ci=r[c_qty] if c_qty is not None else None,
            kiymet_usd=r[c_total] if c_total is not None else None,
            gtip_ci=gtip))
    # döviz
    doviz='USD'
    for i in range(len(raw)):
        t=' '.join(str(v) for v in raw.iloc[i] if pd.notna(v))
        m=re.search(r'Currency[:\s]+([A-Z]{3})',t)
        if m: doviz=m.group(1);break
    return pd.DataFrame(rows), doviz

# ════════════════════ PL OKUMA ════════════════════
def read_pl(path):
    raw=read_any(path)
    hr=None
    for i in range(min(30,len(raw))):
        t=' '.join(str(v).lower() for v in raw.iloc[i] if pd.notna(v))
        if ('description' in t or 'ürün' in t or 'model' in t) and ('qty' in t or 'quantity' in t or 'adet' in t):
            hr=i;break
    if hr is None: raise SystemExit("HATA: PL başlık satırı bulunamadı.")
    hdr=[str(v).strip().lower() if pd.notna(v) else f'c{j}' for j,v in enumerate(raw.iloc[hr])]
    def col(keys):
        for j,h in enumerate(hdr):
            if any(k in h for k in keys): return j
        return None
    c_desc=col(['description','ürün','goods']); c_qty=col(['total qty','quantity','qty','adet','pcs'])
    c_net=col(['n.w','net']); c_gross=col(['g.w','gross','brüt','brut'])
    rows=[]; last=None; ln=0
    for i in range(hr+1,len(raw)):
        r=raw.iloc[i]
        qty=r[c_qty] if c_qty is not None else None
        if pd.isna(qty):
            d=norm(r[c_desc]) if c_desc is not None else None
            if d and 'total' in d.lower(): break
            continue
        d=norm(r[c_desc]) if c_desc is not None else None
        if d: last=d
        else: d=last
        ln+=1
        rows.append(dict(pl_satir=ln, urun=d, adet_pl=qty,
            net_kg=r[c_net] if c_net is not None else None,
            brut_kg=r[c_gross] if c_gross is not None else None))
    return pd.DataFrame(rows)

# ════════════════════ ÖĞRENME OKUMA ════════════════════
def load_learned():
    if os.path.exists(LEARN_XLSX):
        return pd.read_excel(LEARN_XLSX)
    return pd.DataFrame(columns=['Tedarikçi','Ürün Adı (EN)','Model','GTİP','Türkçe Tanım',
        'Tür','Menşe Kodu','KDV %','Ölçü Birimi','Birim Fiyat (örnek)','Kez',
        'İlk Görülme','Son Görülme','Kaynak Dosya'])

def lookup_learned(learned, tedarikci, urun):
    if learned.empty: return None
    m=(learned['Tedarikçi'].map(normU)==normU(tedarikci)) & \
      (learned['Ürün Adı (EN)'].map(normU)==normU(urun))
    if m.any(): return learned[m].iloc[0].to_dict()
    return None

# ════════════════════ ARŞİV ÇAPRAZ KONTROL ════════════════════
def load_ref():
    if os.path.exists(REF_PKL):
        return pickle.load(open(REF_PKL,'rb'))
    return None

def archive_check(ref, urun, gtip_ci):
    if ref is None: return 'arşiv yok'
    u=normU(urun) or ''
    qwords=set(w for w in u.split() if len(w)>=3)
    cands=set()
    for w in qwords: cands|=ref['tanim_words'].get(w,set())
    if not cands: return 'arşivde yok'
    best=None;bs=0
    for c in cands:
        s=SequenceMatcher(None,u,c).ratio()
        if s>bs: bs=s;best=c
    if best and best in ref['lut_tanim']:
        ga=str(ref['lut_tanim'][best]['gtip']).replace('.','')
        gc=str(gtip_ci).replace('.','') if gtip_ci else ''
        if gc and ga==gc: return 'ARŞİV TUTUYOR'
        elif gc: return f'DİKKAT: arşiv {dot_gtip(ga)}'
        else: return f'arşiv önerisi {dot_gtip(ga)}'
    return 'arşivde yok'

# ════════════════════ ANA İŞLEM ════════════════════
def process(ci_path, pl_path, dosya, tedarikci=None, mense=None):
    ci, doviz = read_ci(ci_path)
    pl = read_pl(pl_path)
    ref = load_ref()
    learned = load_learned()
    if not tedarikci:
        # CI'nın ilk satırından tedarikçi tahmini zor; parametre öner
        tedarikci = 'BİLİNMİYOR'

    pl_pool = pl.copy()
    detay=[]
    for _,c in ci.iterrows():
        cand=pl_pool[(pl_pool['urun']==c['urun'])&(pl_pool['adet_pl']==c['adet_ci'])]
        if len(cand)==0: cand=pl_pool[pl_pool['adet_pl']==c['adet_ci']]
        if len(cand)==0: cand=pl_pool[pl_pool['urun']==c['urun']]
        gtip=c['gtip_ci']
        # öğrenmeden GTİP tamamla (faturada yoksa)
        kaynak_gtip='CI'
        L=lookup_learned(learned, tedarikci, c['urun'])
        if not gtip and L:
            gtip=str(L['GTİP']); kaynak_gtip='ÖĞRENİLMİŞ'
        elif not gtip:
            ac=archive_check(ref,c['urun'],None)
            m=re.search(r'(\d{10,12})',ac.replace('.',''))
            if m: gtip=m.group(1); kaynak_gtip='ARŞİV'
        tr_tanim = L['Türkçe Tanım'] if L else ''
        kdv = L['KDV %'] if L else ''
        row=dict(ci_satir=c['ci_satir'], urun=c['urun'], model=c['model'],
                 gtip=gtip, gtip_kaynak=kaynak_gtip, tur='Bedelli',
                 adet_ci=c['adet_ci'], birim_fiyat=c['birim_fiyat'], kiymet_usd=c['kiymet_usd'],
                 tr_tanim=tr_tanim, kdv=kdv)
        if len(cand)>0:
            p=cand.iloc[0]; pl_pool=pl_pool.drop(p.name)
            row.update(pl_satir=p['pl_satir'], adet_pl=p['adet_pl'],
                       net_kg=p['net_kg'], brut_kg=p['brut_kg'],
                       adet_durum='OK' if c['adet_ci']==p['adet_pl'] else 'FARK',
                       isim_durum='OK' if c['urun']==p['urun'] else 'FARK')
        else:
            row.update(pl_satir=None, adet_pl=None, net_kg=None, brut_kg=None,
                       adet_durum='PL YOK', isim_durum='-')
        row['arsiv']=archive_check(ref,c['urun'],gtip)
        detay.append(row)
    D=pd.DataFrame(detay)
    return D, doviz, tedarikci, ref, learned

# ════════════════════ EXCEL ÜRET ════════════════════
def build_excel(D, doviz, dosya, tedarikci):
    from openpyxl import Workbook
    from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
    from openpyxl.utils import get_column_letter
    H=PatternFill("solid",fgColor="1F3864");OK=PatternFill("solid",fgColor="C6EFCE")
    WARN=PatternFill("solid",fgColor="FFC7CE");DIK=PatternFill("solid",fgColor="FFEB9C")
    HDRB=PatternFill("solid",fgColor="D9E1F2")
    thin=Side(style="thin",color="BFBFBF");BD=Border(thin,thin,thin,thin)
    AR=Font(name="Arial",size=10); ARB=Font(name="Arial",size=10,bold=True)
    def hdr(ws,n):
        for j in range(1,n+1):
            c=ws.cell(1,j);c.fill=H;c.font=Font(name="Arial",bold=True,color="FFFFFF",size=10)
            c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True);c.border=BD
        ws.row_dimensions[1].height=32;ws.freeze_panes="A2"
    wb=Workbook()

    # S1 Detay
    ws=wb.active;ws.title="Detay_Eslestirme"
    cols=['CI Satır','PL Satır','GTİP','GTİP Kaynağı','Ürün','Türkçe Tanım','Tür','Adet CI','Adet PL',
          'Birim Fiyat','Kıymet USD','Net Kilo','Brüt Kilo','KDV %','Adet Durumu','İsim Durumu','Arşiv Kontrol']
    ws.append(cols)
    for _,r in D.iterrows():
        ws.append([r['ci_satir'],r['pl_satir'],dot_gtip(r['gtip']) if r['gtip'] else '',
            r['gtip_kaynak'],r['urun'],r['tr_tanim'],r['tur'],r['adet_ci'],r['adet_pl'],
            r['birim_fiyat'],r['kiymet_usd'],r['net_kg'],r['brut_kg'],r['kdv'],
            r['adet_durum'],r['isim_durum'],r['arsiv']])
    hdr(ws,len(cols))
    ia=cols.index('Adet Durumu')+1;ii=cols.index('İsim Durumu')+1;iar=cols.index('Arşiv Kontrol')+1
    ig=cols.index('GTİP Kaynağı')+1
    for i in range(2,ws.max_row+1):
        for j in range(1,len(cols)+1):
            c=ws.cell(i,j);c.border=BD;c.font=AR
            c.alignment=Alignment(vertical="center",wrap_text=(cols[j-1] in('Ürün','Türkçe Tanım','Arşiv Kontrol')))
        for ci_ in (ia,ii):
            ws.cell(i,ci_).fill = OK if ws.cell(i,ci_).value=='OK' else WARN
        av=str(ws.cell(i,iar).value)
        ws.cell(i,iar).fill = OK if 'TUTUYOR' in av else (DIK if 'DİKKAT' in av else PatternFill())
        gk=str(ws.cell(i,ig).value)
        ws.cell(i,ig).fill = OK if gk in('CI','ÖĞRENİLMİŞ') else (DIK if gk in('ARŞİV',) else WARN if gk=='' else PatternFill())
    for j,w in enumerate([7,7,17,12,24,22,8,9,9,10,11,9,9,7,11,11,20],1):
        ws.column_dimensions[get_column_letter(j)].width=w

    # S2 Özet
    ws2=wb.create_sheet("Tarife_Bazinda_Ozet")
    Dg=D.dropna(subset=['gtip']).copy()
    Dg['gtip']=Dg['gtip'].astype(str)
    grp=Dg.groupby(['gtip','tur']).agg(
        ds=('ci_satir','count'),ad=('adet_ci','sum'),ky=('kiymet_usd','sum'),
        nk=('net_kg','sum'),bk=('brut_kg','sum'),
        ur=('urun',lambda x:' / '.join(sorted(set(x)))),
        tr=('tr_tanim',lambda x:next((v for v in x if v),''))).reset_index()
    oc=['GTİP','Türkçe Tanım','Tür','Ürün İsimleri','Detay Satır','Adet','Kıymet USD','Net Kilo','Brüt Kilo']
    ws2.append(oc)
    for _,r in grp.iterrows():
        ws2.append([dot_gtip(r['gtip']),r['tr'],r['tur'],r['ur'],int(r['ds']),
            r['ad'],round(r['ky'],2),round(r['nk'],2),round(r['bk'],3)])
    tr=ws2.max_row+1;ws2.cell(tr,1,'TOPLAM')
    for col_l in ['E','F','G','H','I']: ws2.cell(tr,ord(col_l)-64,f"=SUM({col_l}2:{col_l}{tr-1})")
    hdr(ws2,len(oc))
    for i in range(2,ws2.max_row+1):
        for j in range(1,len(oc)+1):
            c=ws2.cell(i,j);c.border=BD;c.font=ARB if i==tr else AR
            c.alignment=Alignment(vertical="center",wrap_text=(oc[j-1] in('Ürün İsimleri','Türkçe Tanım')))
            if i==tr:c.fill=HDRB
    for j,w in enumerate([17,22,8,40,10,11,12,10,10],1):
        ws2.column_dimensions[get_column_letter(j)].width=w

    # S3 Kontrol
    ws3=wb.create_sheet("Kontrol")
    ci_ad=D['adet_ci'].sum();pl_ad=D['adet_pl'].sum();ci_ky=D['kiymet_usd'].sum()
    oz_ky=grp['ky'].sum();oz_nk=grp['nk'].sum();oz_bk=grp['bk'].sum()
    isimf=(D['isim_durum']=='FARK').sum();adetf=(D['adet_durum']!='OK').sum()
    dik=(D['arsiv'].astype(str).str.contains('DİKKAT')).sum()
    nogtip=(D['gtip'].isna()|(D['gtip']=='')).sum()
    K=[('Kontrol Kalemi','Değer','Durum'),
       ('CI Toplam Adet',f"{ci_ad:,.0f}",''),('PL Toplam Adet',f"{pl_ad:,.0f}",''),
       ('Adet Farkı',f"{ci_ad-pl_ad:,.0f}",'OK' if ci_ad==pl_ad else 'HATA'),
       ('CI Toplam Kıymet',f"{ci_ky:,.2f} {doviz}",''),
       ('Özet Kıymet',f"{oz_ky:,.2f} {doviz}",'OK' if abs(ci_ky-oz_ky)<0.01 else 'HATA'),
       ('Özet Net Kilo',f"{oz_nk:,.2f}",''),('Özet Brüt Kilo',f"{oz_bk:,.2f}",''),
       ('GTİP Sayısı',f"{grp.shape[0]}",''),
       ('— HATA YAKALAMA —','',''),
       ('İsim Farkı Olan Satır',f"{isimf}",'OK' if isimf==0 else 'İNCELE'),
       ('Adet Uyuşmayan Satır',f"{adetf}",'OK' if adetf==0 else 'İNCELE'),
       ('GTİP Boş Kalem',f"{nogtip}",'OK' if nogtip==0 else 'İNCELE'),
       ('Arşiv GTİP Uyuşmazlığı',f"{dik}",'OK' if dik==0 else 'İNCELE')]
    for row in K: ws3.append(row)
    hdr(ws3,3)
    for i in range(2,ws3.max_row+1):
        for j in range(1,4):
            c=ws3.cell(i,j);c.border=BD;c.font=AR;c.alignment=Alignment(vertical="center")
        st=ws3.cell(i,3).value
        if st=='OK':ws3.cell(i,3).fill=OK
        elif st=='HATA':ws3.cell(i,3).fill=WARN
        elif st=='İNCELE':ws3.cell(i,3).fill=DIK
        if ws3.cell(i,1).value and '—' in str(ws3.cell(i,1).value):
            for j in range(1,4):ws3.cell(i,j).fill=HDRB;ws3.cell(i,j).font=ARB
    for j,w in enumerate([26,18,12],1):ws3.column_dimensions[get_column_letter(j)].width=w

    # S4 XML Hazirlik
    ws4=wb.create_sheet("XML_Hazirlik")
    X=[('XML Alanı','Kaynak','Kural','Durum'),
       ('Gtip','Özet','GTİP + tür bazında','Hazır'),
       ('Miktar / Adedi','Özet','Adet','Hazır'),
       ('Birim_fiyat','Özet','Kıymet / Adet','Hazır'),
       ('Brut_agirlik','Özet','PL brüt kilo','Hazır'),
       ('Net_agirlik','Özet','PL net kilo','Hazır'),
       ('Ticari_tanimi','Özet','Türkçe tanım','SEN ONAYLA'),
       ('Mensei_ulke','Sen gir/öğrenme','Menşe kodu','SEN GİR'),
       ('Doviz','CI',doviz,'Hazır'),
       ('Kdv_orani','Öğrenme','GTİP bazında','SEN ONAYLA'),
       ('Beyanname başlığı','Sen gir','Rejim/gümrük/teslim/navlun/sigorta','SEN GİR')]
    for row in X: ws4.append(row)
    hdr(ws4,4)
    for i in range(2,ws4.max_row+1):
        for j in range(1,5):
            c=ws4.cell(i,j);c.border=BD;c.font=AR;c.alignment=Alignment(vertical="center",wrap_text=True)
        st=ws4.cell(i,4).value
        if st=='Hazır':ws4.cell(i,4).fill=OK
        elif 'SEN' in str(st):ws4.cell(i,4).fill=DIK
    for j,w in enumerate([24,18,38,12],1):ws4.column_dimensions[get_column_letter(j)].width=w

    os.makedirs(CIKTI_DIR,exist_ok=True)
    out=os.path.join(CIKTI_DIR,f"{dosya}_HAZIRLIK.xlsx")
    wb.save(out)
    return out, dict(isim_fark=isimf,adet_fark=adetf,nogtip=nogtip,dikkat=dik,
                     kiymet_ok=abs(ci_ky-oz_ky)<0.01)

# ════════════════════ ÖĞRENME İŞLE ════════════════════
def learn(D, tedarikci, dosya, mense='', kdv='', olcu='ADET'):
    learned=load_learned()
    bugun=datetime.now().strftime('%Y-%m-%d')
    add=upd=0
    for _,r in D.iterrows():
        urun=norm(r['urun']);gtip=str(r['gtip']).replace('.','').replace(' ','') if r['gtip'] else None
        if not urun or not gtip: continue
        m=(learned['Tedarikçi'].map(normU)==normU(tedarikci))&(learned['Ürün Adı (EN)'].map(normU)==normU(urun))
        if m.any():
            i=learned[m].index[0]
            learned.at[i,'Kez']=int(learned.at[i,'Kez'])+1
            learned.at[i,'Son Görülme']=bugun
            learned.at[i,'GTİP']=gtip
            if r['tr_tanim']: learned.at[i,'Türkçe Tanım']=r['tr_tanim']
            upd+=1
        else:
            learned=pd.concat([learned,pd.DataFrame([{
                'Tedarikçi':norm(tedarikci),'Ürün Adı (EN)':urun,'Model':norm(r['model']) or '',
                'GTİP':gtip,'Türkçe Tanım':r['tr_tanim'] or '','Tür':r['tur'],
                'Menşe Kodu':mense,'KDV %':r['kdv'] or kdv,'Ölçü Birimi':olcu,
                'Birim Fiyat (örnek)':r['birim_fiyat'],'Kez':1,
                'İlk Görülme':bugun,'Son Görülme':bugun,'Kaynak Dosya':dosya}])],ignore_index=True)
            add+=1
    os.makedirs(os.path.dirname(LEARN_XLSX),exist_ok=True)
    learned.to_excel(LEARN_XLSX,index=False)
    return add,upd,len(learned)

# ════════════════════ ÇAKIŞMA KORUMASI ════════════════════
def guvenli_yol(dir_, base_ad):
    """Aynı isimde dosya varsa _2, _3 ekle — üstüne yazma."""
    yol=os.path.join(dir_, base_ad+'.xlsx')
    if not os.path.exists(yol): return yol
    n=2
    while True:
        yol=os.path.join(dir_, f"{base_ad}_{n}.xlsx")
        if not os.path.exists(yol): return yol
        n+=1

# ════════════════════ KLASÖR BULMA ════════════════════
def bul_ci_pl(klasor):
    """Bir klasörde CI ve PL dosyalarını isimden ayırt et."""
    dosyalar=[f for f in os.listdir(klasor) if f.lower().endswith(('.xls','.xlsx')) and not f.startswith('~')]
    ci=pl=None
    for f in dosyalar:
        fl=f.lower()
        if re.search(r'\bci\b|invoice|fatura|_ci_|ci_for|ci for', fl): ci=os.path.join(klasor,f)
        elif re.search(r'\bpl\b|packing|çeki|ceki|_pl_|pl_for|pl for', fl): pl=os.path.join(klasor,f)
    # tek dosyada birleşik CI+PL olabilir
    if not ci and not pl and len(dosyalar)==1:
        ci=pl=os.path.join(klasor,dosyalar[0])
    # ayırt edilemezse: 2 dosya varsa ilki CI ikincisi PL varsay
    if not ci and len(dosyalar)>=1: ci=os.path.join(klasor,sorted(dosyalar)[0])
    if not pl and len(dosyalar)>=2: pl=os.path.join(klasor,sorted(dosyalar)[1])
    if not pl: pl=ci
    return ci, pl

# ════════════════════ MAIN ════════════════════
def main():
    ap=argparse.ArgumentParser(description="Gümrük beyanname hazırlık")
    ap.add_argument('--klasor', help='gelen/ altındaki iş klasörü (örn: gelen/GODSON1). Verilmezse gelen/ altındaki ilk klasör.')
    ap.add_argument('--ci', help='CI dosyası (klasör yerine tek tek de verilebilir)')
    ap.add_argument('--pl', help='PL dosyası')
    ap.add_argument('--tedarikci', default=None, help='Otomatik bulunur; elle geçmek istersen')
    ap.add_argument('--mense', default=None, help='Menşe kodu (otomatik/varsayılan 720)')
    ap.add_argument('--dosya', default=None, help='Çıktı adı (otomatik: FIRMA_FATURANO)')
    ap.add_argument('--onayla', action='store_true', help='Onayla ve öğren + tamamlanana taşı')
    a=ap.parse_args()

    GELEN=os.path.join(BASE,'gelen'); TAMAM=os.path.join(BASE,'tamamlanan')

    # --- CI/PL yolunu belirle ---
    is_klasoru=None
    if a.ci:
        ci_path=a.ci; pl_path=a.pl or a.ci
    else:
        if a.klasor:
            is_klasoru = a.klasor if os.path.isdir(a.klasor) else os.path.join(GELEN, a.klasor)
        else:
            # gelen/ altındaki ilk alt klasör; yoksa gelen/'in kendisi
            alt=[d for d in (os.listdir(GELEN) if os.path.isdir(GELEN) else [])
                 if os.path.isdir(os.path.join(GELEN,d)) and d!='tamamlanan']
            is_klasoru = os.path.join(GELEN, sorted(alt)[0]) if alt else GELEN
        if not os.path.isdir(is_klasoru):
            raise SystemExit(f"HATA: klasör yok: {is_klasoru}")
        ci_path, pl_path = bul_ci_pl(is_klasoru)
        if not ci_path: raise SystemExit(f"HATA: {is_klasoru} içinde CI/PL bulunamadı.")

    # --- Otomatik meta ---
    meta=extract_meta(ci_path)
    tedarikci = a.tedarikci or meta['tedarikci'] or 'BİLİNMİYOR'
    mense = a.mense or meta['mense_kod']
    fatno = meta['fatno'] or 'NOFAT'
    dosya = a.dosya or f"{kisa_ad(tedarikci)}_{fatno}"

    print(f"İşleniyor: {dosya}")
    print(f"  Tedarikçi (otomatik): {tedarikci}")
    print(f"  Fatura No: {fatno} | Menşe: {mense}")
    if is_klasoru: print(f"  Kaynak klasör: {is_klasoru}")

    D,doviz,ted,ref,learned=process(ci_path,pl_path,dosya,tedarikci,mense)
    ted=tedarikci
    print(f"  CI+PL eşleşti: {len(D)} kalem | döviz {doviz}")

    out,ozet=build_excel(D,doviz,dosya,ted)
    # çakışma koruması: build_excel sabit isim yazdı, güvenli isme taşı
    hedef=guvenli_yol(CIKTI_DIR, f"{dosya}_HAZIRLIK")
    if os.path.abspath(out)!=os.path.abspath(hedef):
        os.replace(out,hedef); out=hedef
    print(f"  Çıktı: {out}")
    print(f"  Hata yakalama: isim farkı={ozet['isim_fark']}, adet fark={ozet['adet_fark']}, "
          f"GTİP boş={ozet['nogtip']}, arşiv uyuşmazlık={ozet['dikkat']}, "
          f"kıymet mutabakat={'OK' if ozet['kiymet_ok'] else 'HATA'}")

    if a.onayla:
        add,upd,tot=learn(D,ted,dosya,mense)
        print(f"  ÖĞRENİLDİ: +{add} yeni, {upd} güncelleme, toplam {tot} kayıt")
        # tamamlanana taşı: CI/PL + çıktı bir arada
        if is_klasoru and is_klasoru!=GELEN:
            os.makedirs(TAMAM,exist_ok=True)
            hedef_kl=os.path.join(TAMAM, os.path.basename(is_klasoru))
            try:
                import shutil
                shutil.copy(out, is_klasoru)  # çıktıyı iş klasörüne de koy
                os.replace(is_klasoru, hedef_kl)
                print(f"  TAMAMLANANA TAŞINDI: {hedef_kl}")
            except Exception as e:
                print(f"  (taşıma atlandı: {e})")
    else:
        print("  (Öğrenme/taşıma yapılmadı. Çıktıyı kontrol edip onayladıktan sonra --onayla ile tekrar çalıştır.)")

if __name__=='__main__':
    main()
